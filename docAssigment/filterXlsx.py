import openpyxl


def removeLine(sheet,keepId,colNb,min_row):
    for row in sheet.iter_rows(min_row=min_row):
        if not sheet.cell(row=row[0].row,column=colNb).value in keepId:
            sheet.delete_rows(row[0].row,1)
            removeLine(sheet,keepId,colNb,min_row=row[0].row)
            return

def filterXlsx(inputFile:str, sheetName:str, colId:str, keepId:list, outputFile:str):
    wb=openpyxl.load_workbook(inputFile)
    sheet=wb[sheetName]
    columns=[sheet.cell(row=1,column=i).value for i in range(1,sheet.max_column)]
    idCol=columns.index(colId)+1
    #for row in sheet:
    removeLine(sheet,keepId,idCol,min_row=2)
    wb.save(outputFile)

# Example:
# filterXlsx('../../screening/fullTextScreening/includedFullTextScreening.xlsx','Sheet 1', 'abbrev', ['Supelano2022','Yarker2013'],'../../screening/fullTextScreening/testFilter2.xlsx')
